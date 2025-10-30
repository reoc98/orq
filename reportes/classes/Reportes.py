from models.RequestsOrquestador import RequestsOrquestador as RequestsOrquestadorModel 
from models.EjecucionMotor import EjecucionMotor as EjecucionMotorModel
from models.Arbol import Arbol as ArbolMotorModel
from models.ReportesSiniestros import ReportesSiniestros
from models.Departamentos import Departamentos
from models.Meses import Meses
from models.UsuariosModel import UsuariosModel
from models.ReporteMensualSiniestros import ReporteMensualSiniestros
from models.EstadosReportes import EstadosReportes
from classes.Database import Database
from classes.PermissionsDatabase import PermissionsDatabase
from utils.Aws import Aws, os
from datetime import datetime, date
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.styles import Font
from openpyxl.styles import Border
from openpyxl.styles import Side
import json
from sqlalchemy import func, asc, and_, or_
import calendar

class Reportes():
    def __init__(self, payload = {}):
        self.payload = payload
    
    # Obtener request de orquestador por id
    def obtenerRequestOrqID(self, id):
        db = Database('dbr')
        try:
            isset = db.session.query(RequestsOrquestadorModel).filter_by(
                id = id).filter_by(ejecucion = 4).filter_by(risk = 1).filter_by(
                estado = 1).filter((RequestsOrquestadorModel.id_estados_reportes == 1) | (RequestsOrquestadorModel.id_estados_reportes == 4)).first()
            # se valida si hay registro
            if(isset == None):
                return None
            
            return isset
        finally:
            db.session.invalidate()
            db.session.close()

    # validamos en la bd si hay un reporte procesandose
    def obtenerReporteOrqEstado(self):
        db = Database('dbr')
        try:
            isset = db.session.query(ReportesSiniestros).filter_by(
                estado = 1).filter((ReportesSiniestros.id_estados_reportes == 3) | (ReportesSiniestros.id_estados_reportes == 2)).first()
            # se valida si hay registro
            if(isset == None):
                return None
            
            return isset
        finally:
            db.session.invalidate()
            db.session.close()
    
    # obtenemos todos los reportes de siniestro pendientes
    def obtenerReporteOrqPendientes(self, estado):
        db = Database('dbr')
        try:
            isset = db.session.query(ReportesSiniestros).filter_by(
                estado = 1).filter_by(id_estados_reportes = estado).all()
            
            return isset
        finally:
            db.session.invalidate()
            db.session.close()

    # obtenemos reporte siniestro por id
    def obtenerReporteOrqId(self, id):
        db = Database('dbr')
        try:
            isset = db.session.query(ReportesSiniestros).filter_by(
                estado = 1).filter((ReportesSiniestros.id == id)).all()
            # se valida si hay registro
            if(isset == None):
                return None
            
            return isset
        finally:
            db.session.invalidate()
            db.session.close()
    
    def get_num_siniestro(self, id_request):
        db = Database('dbr').session
        try:
            row = db.query(RequestsOrquestadorModel.numero_siniestro).filter(
                (RequestsOrquestadorModel.id == id_request)
                ).first()

            if row is None:
                return row
            else:
                return row.numero_siniestro
        
        except Exception as e:
            raise Warning("Error al obtener los datos del siniestro.")
        
        finally:
            db.invalidate()
            db.close()
        
    # obtenemos reporte siniestro por id request y obtenemos la url
    def obtenerReporteOrqIdUrl(self, id_request, id_usuario):

        resp_permission = self.validate_permission(user = id_usuario, type = "LISTAR", module = 8)
        if resp_permission == False:
            raise ValueError("No cuenta con los permisos necesarios para realizar esta operación.")

        db = Database('dbr')
        try:
            isset = db.session.query(ReportesSiniestros).filter_by(
                estado = 1).filter_by(
                id_estados_reportes = 5).filter((ReportesSiniestros.id_request == id_request)).first()
            # se valida si hay registro
            if(isset == None):
                return None
            
            return isset
        finally:
            db.session.invalidate()
            db.session.close()

    # guardamos el reporte en la tabla
    def guardarReporte(self, id_request, estado):
        db = Database('dbw').session
        try:
            # desactivamos el ultimo registro si tiene
            query = db.query(ReportesSiniestros).filter(ReportesSiniestros.id_request == id_request)
            data = {
                'estado': 0
            }
            query.update(data)
            db.commit()

            # guardamos como pendiente en la tabla de reportes siniestro
            row = ReportesSiniestros(id_request, estado)
            db.add(row)
            db.commit()

            # cambiamos el estado cabecera request orquestador
            query = db.query(RequestsOrquestadorModel).filter(RequestsOrquestadorModel.id == id_request)
            data = {
                'id_estados_reportes': estado
            }
            query.update(data)
            db.commit()

            return True
        except Exception as error:
            return False
        
        finally:
            db.invalidate()
            db.close()
    
    # ejecutar reporte de siniestro
    def generarReporteSiniestro(self, id_usuario):

        db = Database('dbr').session
        try:
            usuario = db.query(
                UsuariosModel.id,
                UsuariosModel.email
            ).filter(
                UsuariosModel.id == id_usuario
            ).first()

            data = {
                'id':"",
                'email': usuario.email,
                'id_usuario': id_usuario
            }
            # llamamos la lambda asyncrona
            Aws.lambdaInvoke(
                    str(os.getenv('SERVICE'))
                    + '-'+str(os.getenv('STAGE'))
                    + '-generarReporteSiniestro', data)
        except Exception as error:
            return None
        finally:
            db.invalidate()
            db.close()

    # ejecutar reporte de siniestro
    def generarReporteMensual(self, data):

        try:
            # llamamos la lambda asyncrona
            Aws.lambdaInvoke(
                    str(os.getenv('SERVICE'))
                    + '-'+str(os.getenv('STAGE'))
                    + '-asyncGenerarReporteMensual', data)
        except Exception as error:
            return None

    # actualizar reportes de siniestro
    def actualizarReporte(self, id, id_request, estado, response = None):
        db = Database('dbw').session
        try:
            # cambiamos el estado del reporte
            siniestro = db.query(ReportesSiniestros).filter(ReportesSiniestros.id == id)
            ##Bincula los nuevos datos que se van actualizar
            data = {
                'id_estados_reportes': estado,
                'response': json.dumps(response)
            }
            siniestro.update(data)
            db.commit()

            # cambiamos el estado cabecera
            query = db.query(RequestsOrquestadorModel).filter(RequestsOrquestadorModel.id == id_request)
            data1 = {
                'id_estados_reportes': estado
            }
            query.update(data1)
            db.commit()

            return True
        except Exception as error:
            return False
        
        finally:
            db.invalidate()
            db.close()

    # genera el excel de reporte
    def excelReporteSiniestro(self, siniestro, email, num_siniestro):
        session = Database('dbw').session
        try:

            date1 = datetime.today()
            date = date1.strftime('%Y%m%d%H%M%S')
            file_name = f"Reporte_Siniestro_{date}.xlsx"
            tmp_dir = os.getenv('TMP')
            tmp_path = f"{tmp_dir}{file_name}"
            file_path = f"reporte_siniestro/{file_name}"
            
            wb = Workbook()
            worksheet = wb.active

            f_hoy = date1.strftime('%d/%m/%Y')
            worksheet['A1'] = "Orquestador"
            worksheet['A2'] = "Reporte de Siniestros"
            worksheet['A3'] = f"Reporte generado el {f_hoy} por - {email}"
            worksheet['A4'] = f"Número siniestro: {num_siniestro}"

            worksheet["A1"].font = Font(bold=True)
            worksheet["A2"].font = Font(bold=True)
            worksheet["A3"].font = Font(bold=True)
            worksheet["A4"].font = Font(bold=True)

            worksheet['A6'] = 'CODIGO MOTOR'
            worksheet["A6"].font = Font(bold=True)
            worksheet['B6'] = 'NOMBRE MOTOR'
            worksheet["B6"].font = Font(bold=True)
            worksheet['C6'] = 'TIPO NODO'
            worksheet["C6"].font = Font(bold=True)
            worksheet['D6'] = 'REGLAS EVALUADAS'
            worksheet["D6"].font = Font(bold=True)
            worksheet['E6'] = 'VALOR ESPERADO'
            worksheet["E6"].font = Font(bold=True)
            worksheet['F6'] = 'VALOR RECIBIDO'
            worksheet["F6"].font = Font(bold=True)
            worksheet['G6'] = 'RESULTADO'
            worksheet["G6"].font = Font(bold=True)
            worksheet['H6'] = 'FECHA EJECUCIÓN'
            worksheet["H6"].font = Font(bold=True)
            worksheet.column_dimensions['A'].width = 15
            worksheet.column_dimensions['B'].width = 35
            worksheet.column_dimensions['C'].width = 13
            worksheet.column_dimensions['D'].width = 40
            worksheet.column_dimensions['E'].width = 15
            worksheet.column_dimensions['F'].width = 15
            worksheet.column_dimensions['G'].width = 15
            worksheet.column_dimensions['H'].width = 18
            
            result = session.query(
                EjecucionMotorModel.id,
                EjecucionMotorModel.codigo_arbol,
                EjecucionMotorModel.datos_response,
                EjecucionMotorModel.codigo_arbol,
                EjecucionMotorModel.fecha_upd,
                ArbolMotorModel.nombre
            ).join(
                ArbolMotorModel,
                EjecucionMotorModel.id_arbol == ArbolMotorModel.id
            ).filter(
                EjecucionMotorModel.id_request == siniestro.id_request,
                EjecucionMotorModel.estado == 1
            ).all()

            indx = 7
            for i in result:
                indx_i = indx
                worksheet[f'A{indx}'] = i.codigo_arbol
                worksheet[f'B{indx}'] = i.nombre
                datos_response = json.loads(i.datos_response)

                try:
                    resultado_nodos = datos_response["datos"]["resp_motor"]["resultado_nodos"]
                except:
                    resultado_nodos = []

                for j in resultado_nodos:
                    worksheet[f'C{indx}'] = j["tipo_nodo"]
                    worksheet[f'D{indx}'] = j["regla_evaluada"]
                    worksheet[f'E{indx}'] = j["valor_esperado"]
                    worksheet[f'F{indx}'] = j["valor_recibido"]

                    resultado = "NO APROBADO"
                    if str(j["resultado"]) == "1":
                        resultado = "APROBADO"

                    worksheet[f'G{indx}'] = resultado
                    worksheet[f'H{indx}'] = str(i.fecha_upd)
                    indx += 1
                    nodos = j.get("nodos", "")
                    if nodos != "":
                        for k in j["nodos"]:
                            worksheet[f'C{indx}'] = k["tipo_nodo"]
                            worksheet[f'D{indx}'] = k["regla_evaluada"]
                            worksheet[f'E{indx}'] = k["valor_esperado"]
                            worksheet[f'F{indx}'] = k["valor_recibido"]

                            resultado = "NO APROBADO"
                            if str(k["resultado"]) == "1":
                                resultado = "APROBADO"

                            worksheet[f'G{indx}'] = resultado
                            worksheet[f'H{indx}'] = str(i.fecha_upd)
                            indx += 1

                if indx_i < indx:
                    indx_f = indx - 1
                    worksheet.merge_cells(f'A{indx_i}:A{indx_f}')
                    worksheet.merge_cells(f'B{indx_i}:B{indx_f}')
                else:
                    indx += 1

            x = range(6, indx)
            for n in x:
                worksheet[f"A{n}"].alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                    wrap_text=True
                )
                worksheet[f"B{n}"].alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                    wrap_text=True
                )
                worksheet[f"C{n}"].alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                    wrap_text=True
                )
                worksheet[f"D{n}"].alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                    wrap_text=True
                )
                worksheet[f"E{n}"].alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                    wrap_text=True
                )
                worksheet[f"F{n}"].alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                    wrap_text=True
                )
                worksheet[f"G{n}"].alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                    wrap_text=True
                )
                worksheet[f"H{n}"].alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                    wrap_text=True
                )

                worksheet[f"A{n}"].border = Border(
                    left=Side(border_style='thin'),
                    right=Side(border_style='thin'),
                    top=Side(border_style='thin'),
                    bottom=Side(border_style='thin')
                )
                worksheet[f"B{n}"].border = Border(
                    left=Side(border_style='thin'),
                    right=Side(border_style='thin'),
                    top=Side(border_style='thin'),
                    bottom=Side(border_style='thin')
                )
                worksheet[f"C{n}"].border = Border(
                    left=Side(border_style='thin'),
                    right=Side(border_style='thin'),
                    top=Side(border_style='thin'),
                    bottom=Side(border_style='thin')
                )
                worksheet[f"D{n}"].border = Border(
                    left=Side(border_style='thin'),
                    right=Side(border_style='thin'),
                    top=Side(border_style='thin'),
                    bottom=Side(border_style='thin')
                )
                worksheet[f"E{n}"].border = Border(
                    left=Side(border_style='thin'),
                    right=Side(border_style='thin'),
                    top=Side(border_style='thin'),
                    bottom=Side(border_style='thin')
                )
                worksheet[f"F{n}"].border = Border(
                    left=Side(border_style='thin'),
                    right=Side(border_style='thin'),
                    top=Side(border_style='thin'),
                    bottom=Side(border_style='thin')
                )
                worksheet[f"G{n}"].border = Border(
                    left=Side(border_style='thin'),
                    right=Side(border_style='thin'),
                    top=Side(border_style='thin'),
                    bottom=Side(border_style='thin')
                )
                worksheet[f"H{n}"].border = Border(
                    left=Side(border_style='thin'),
                    right=Side(border_style='thin'),
                    top=Side(border_style='thin'),
                    bottom=Side(border_style='thin')
                )

            wb.save(tmp_path)
            
            secret = Aws.getSecret(str(os.getenv('STAGE'))+"/secret-orquestador")
            Aws.subirArchivoS3(
                secret['bucket_s3'], tmp_path, file_path)

            query = session.query(ReportesSiniestros).filter(ReportesSiniestros.id == siniestro.id)

            ##Vincula los nuevos datos que se van actualizar
            data = {
                'url': file_path
            }
            query.update(data)
            session.commit()

            return True
        
        except ValueError as e:
            raise Exception(f"Error al generar excel "+str(e))
        except Exception as e:
            raise Exception(f"Error al generar excel "+str(e))
        finally:
            session.invalidate()
            session.close()

    def validar_estado_reporte(self, id_request, id_usuario):
        db = Database('dbw').session
        try:
            resp_permission = self.validate_permission(user = id_usuario, type = "LISTAR", module = 8)
            if resp_permission == False:
                raise ValueError("No cuenta con los permisos necesarios para realizar esta operación.")
            
            # cambiamos el estado del reporte
            consult = db.query(RequestsOrquestadorModel.id_estados_reportes, EstadosReportes.nombre.label("tipo_estado")).join(
                EstadosReportes, EstadosReportes.id == RequestsOrquestadorModel.id_estados_reportes
            ).filter(
                RequestsOrquestadorModel.id == id_request
            )
            
            siniestro = consult.first()
            
            if siniestro.id_estados_reportes == 2:
                return True
            
            else:
                raise ValueError(f"El reporte actualmente se encuentra en estado '{siniestro.tipo_estado}'.")

        except Exception as e:
            Type = type(e)
            raise Type(str(e))
        
        finally:
            db.invalidate()
            db.close()

    # cancelamos un reporte en pendiente
    def cancelarReporte(self, id_request):
        db = Database('dbw').session
        try:
            ##Bincula los nuevos datos que se van actualizar
            data = {
                'estado': 0
            }
            # cambiamos el estado del reporte
            siniestro = db.query(ReportesSiniestros).filter(ReportesSiniestros.id_request == id_request).filter(
                ReportesSiniestros.id_estados_reportes == 2).filter(
                ReportesSiniestros.estado == 1).update(data)
            db.commit()

            if siniestro:
                # cambiamos el estado cabecera
                query = db.query(RequestsOrquestadorModel).filter(RequestsOrquestadorModel.id == id_request)
                data1 = {
                    'id_estados_reportes': 1
                }
                query.update(data1)
                db.commit()
            else:
                raise Exception('El reporte no se encuentra en estado pendiente.')

            return True
        except Exception as error:
            return False
        
        finally:
            db.invalidate()
            db.close()

    # funcion que busca la url en el s3 y genera el link de descarga
    def descargarReporte(self, url):

        secret = Aws.getSecret(str(os.getenv('STAGE'))+"/secret-orquestador")

        return Aws.generarLinkDescargaS3(secret['bucket_s3'], url)

    # guardamos el reporte en la tabla
    def insertarReporteMensual(self):
        db = Database('dbw').session
        try:
            # guardamos como pendiente en la tabla de reportes siniestro
            payload = {
                "estado": 0
            }
            row = ReporteMensualSiniestros(payload)
            db.add(row)
            db.commit()
            info_insert = json.loads(str(row))

            return { "statusCode": 200, "data": info_insert }
        except Exception as error:
            return False
        
        finally:
            db.invalidate()
            db.close()

    def validarUsuario(self, id_usuario):
        db = Database('dbr').session
        try:
            resp_permission = self.validate_permission(user = id_usuario, type = "LISTAR", module=8)
            if resp_permission == False:
                raise ValueError("No cuenta con los permisos necesarios para realizar esta operación.")

            usuario = db.query(
                UsuariosModel.id,
                UsuariosModel.email
            ).filter(
                UsuariosModel.id == id_usuario
            ).first()

            if usuario is None:
                return { "statusCode": 400, "message": "Usuario incorrecto" }
            
            return { "statusCode": 200, "message": "OK"}
            
        except ValueError as e:
            Type = type(e)
            raise Type(str(e))
        
        except Exception as e:
            raise Exception(f"Error al generar excel "+str(e))
        finally:
            db.invalidate()
            db.close()

    # obtener informacion del reporte mensual
    def informacionReporte(self, data):
        db = Database('dbr').session
        dbw = Database('dbw').session
        try:
            ##Bincula los nuevos datos que se van actualizar
            
            anio = data["anio"]
            mes = data["mes"]
            departamento = data["departamento"]
            id_usuario = data["id_usuario"]
            id_reporte = data["id_reporte"]

            nombre_mes = db.query(
                Meses.id,
                Meses.descripcion
            ).filter(
                Meses.id == mes
            ).first()

            usuario = db.query(
                UsuariosModel.id,
                UsuariosModel.email
            ).filter(
                UsuariosModel.id == id_usuario
            ).first()

            if mes not in range(1,13):
                return { "statusCode": 400, "message": "Mes invalido"}

            if anio < 2023 or len(str(anio)) > 4:
                return { "statusCode": 400, "message": "Año invalido"}

            num_dias = calendar.monthrange(anio, mes)[1]
            fecha_inicio = date(anio, mes, 1)
            fecha_fin = date(anio, mes, num_dias)

            if departamento == 0:
                query1 = db.query(
                    func.max(RequestsOrquestadorModel.id).label("id"),
                    func.count(RequestsOrquestadorModel.id).label('validations')
                ).filter(
                    RequestsOrquestadorModel.fecha_reg.between(fecha_inicio, fecha_fin),
                    and_(or_(and_(RequestsOrquestadorModel.ejecucion == 4,
                                    RequestsOrquestadorModel.estado == 1),
                                    RequestsOrquestadorModel.estado == 0))
                ).group_by(
                    RequestsOrquestadorModel.numero_siniestro
                ).all()
            else:
                query1 = db.query(
                    func.max(RequestsOrquestadorModel.id).label("id"),
                    func.count(RequestsOrquestadorModel.id).label('validations'),
                    Departamentos.descripcion.label("departamento")
                ).join(
                    Departamentos,
                    and_(RequestsOrquestadorModel.cod_departamento == Departamentos.codigo,
                        Departamentos.id == departamento)
                ).filter(
                    RequestsOrquestadorModel.fecha_reg.between(fecha_inicio, fecha_fin),
                    and_(or_(and_(RequestsOrquestadorModel.ejecucion == 4,
                                    RequestsOrquestadorModel.estado == 1),
                                    RequestsOrquestadorModel.estado == 0))
                ).group_by(
                    RequestsOrquestadorModel.numero_siniestro
                ).all()

            if len(query1) < 1:
                query = dbw.query(
                    ReporteMensualSiniestros
                ).filter(
                    ReporteMensualSiniestros.id == id_reporte
                )

                ##Vincula los nuevos datos que se van actualizar
                payload = {
                    'estado': 2
                }
                query.update(payload)
                dbw.commit()
                msj = "No se encontraron registros"
                return { "statusCode": 200, "message": msj, "data": [] }

            siniestros = []
            estadisticas = []
            cont_fallidos = 0
            cont_una_vez = 0
            cont_varias_veces = 0
            suma_siniestros = 0
            for i in query1:
                query = db.query(
                    RequestsOrquestadorModel.id,
                    RequestsOrquestadorModel.numero_siniestro,
                    Departamentos.descripcion.label("departamento"),
                    RequestsOrquestadorModel.estado,
                    RequestsOrquestadorModel.ejecucion,
                    RequestsOrquestadorModel.fecha_reg
                ).join(
                    Departamentos,
                    RequestsOrquestadorModel.cod_departamento == Departamentos.codigo,
                    isouter=True
                ).filter(
                    RequestsOrquestadorModel.id == i.id
                ).group_by(
                    RequestsOrquestadorModel.numero_siniestro
                ).first()

                dic = {
                    "id": query.id,
                    "numero_siniestro": query.numero_siniestro,
                    "departamento": query.departamento,
                    "estado": query.estado,
                    "ejecucion": query.ejecucion,
                    "validaciones": i.validations,
                    "fecha_reg": str(query.fecha_reg)
                }

                if query.estado == 0:
                    cont_fallidos += 1

                if i.validations == 1 and query.estado == 1:
                    cont_una_vez += 1

                if i.validations > 1 and query.estado == 1:
                    cont_varias_veces += 1

                suma_siniestros += 1

                siniestros.append(dic)

                nombre_departamento = "Todos"
                if departamento != 0:
                    nombre_departamento = i.departamento
            
            porc_fallidos = round((cont_fallidos/suma_siniestros)*100, 2)
            porc_una_vez = round((cont_una_vez/suma_siniestros)*100, 2)
            porc_varias_veces = round((cont_varias_veces/suma_siniestros)*100, 2)

            estadisticas = {
                "departamento": nombre_departamento,
                "anio": anio,
                "mes": nombre_mes.descripcion,
                "fallidos": cont_fallidos,
                "una_vez": cont_una_vez,
                "varias_veces": cont_varias_veces,
                "porc_fallidos": f"{porc_fallidos}%",
                "porc_una_vez": f"{porc_una_vez}%",
                "porc_varias_veces": f"{porc_varias_veces}%",
                "total": suma_siniestros
            }

            datos = {
                "siniestros": siniestros,
                "estadisticas": estadisticas,
                "fecha_inicio": str(fecha_inicio),
                "fecha_fin": str(fecha_fin),
                "email": usuario.email
            }

            return {"statusCode": 200, "data": datos}
        except Exception as error:
            return {"statusCode": 500, "message": str(error)}
        
        finally:
            db.invalidate()
            db.close()
            dbw.invalidate()
            dbw.close()

    # genera el excel de reporte
    def excelReporteMensualSiniestro(self, datos, id_reporte):
        session = Database('dbw').session
        try:

            date1 = datetime.today()
            date = date1.strftime('%Y%m%d%H%M%S')
            file_name = f"Reporte_Mensual_Siniestro_{date}.xlsx"
            tmp_dir = os.getenv('TMP')
            tmp_path = f"{tmp_dir}{file_name}"
            file_path = f"reporte_siniestro/{file_name}"
            
            wb = Workbook()
            worksheet = wb.active

            fecha_inicio = datos["fecha_inicio"]
            fecha_fin = datos["fecha_fin"]
            email = datos["email"]

            fecha_inicio = datetime.strptime(fecha_inicio, '%Y-%m-%d').date()
            f_inicio = fecha_inicio.strftime('%d/%m/%Y')

            fecha_fin = datetime.strptime(fecha_fin, '%Y-%m-%d').date()
            f_fin = fecha_fin.strftime('%d/%m/%Y')

            f_hoy = date1.strftime('%d/%m/%Y')
            worksheet['A1'] = "Orquestador"
            worksheet['A2'] = "Reporte Mensual de siniestros"
            worksheet['A3'] = f"Reporte generado entre {f_inicio} al {f_fin}"
            worksheet['A4'] = f"Reporte generado el {f_hoy} por - {email}"
            worksheet["A1"].font = Font(bold=True)
            worksheet["A2"].font = Font(bold=True)
            worksheet["A3"].font = Font(bold=True)
            worksheet["A4"].font = Font(bold=True)


            worksheet['A6'] = 'NUM SINIESTRO'
            worksheet["A6"].font = Font(bold=True)
            worksheet['B6'] = 'NUM DE EVALUACIONES'
            worksheet["B6"].font = Font(bold=True)
            worksheet['C6'] = 'FALLIDO'
            worksheet["C6"].font = Font(bold=True)
            worksheet['D6'] = 'DEPARTAMENTO'
            worksheet["D6"].font = Font(bold=True)
            worksheet['E6'] = 'FECHA DE EVALUACIÓN'
            worksheet["E6"].font = Font(bold=True)

            worksheet.column_dimensions['A'].width = 16
            worksheet.column_dimensions['B'].width = 22
            worksheet.column_dimensions['D'].width = 16
            worksheet.column_dimensions['E'].width = 21

            siniestros = datos["siniestros"]
            estadisticas = datos["estadisticas"]

            indx = 7
            for i in siniestros:
                worksheet[f'A{indx}'] = i["numero_siniestro"]
                worksheet[f'B{indx}'] = i["validaciones"]

                fallido = "SI"
                if str(i["estado"]) == "1":
                    fallido = "NO"

                worksheet[f'C{indx}'] = fallido
                worksheet[f'D{indx}'] = i["departamento"]
                worksheet[f'E{indx}'] = i["fecha_reg"]

                indx += 1

            x = range(6, indx)
            for n in x:
                worksheet[f"A{n}"].alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                    wrap_text=True
                )
                worksheet[f"B{n}"].alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                    wrap_text=True
                )
                worksheet[f"C{n}"].alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                    wrap_text=True
                )
                worksheet[f"D{n}"].alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                    wrap_text=True
                )
                worksheet[f"E{n}"].alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                    wrap_text=True
                )

                worksheet[f"A{n}"].border = Border(
                    left=Side(border_style='thin'),
                    right=Side(border_style='thin'),
                    top=Side(border_style='thin'),
                    bottom=Side(border_style='thin')
                )
                worksheet[f"B{n}"].border = Border(
                    left=Side(border_style='thin'),
                    right=Side(border_style='thin'),
                    top=Side(border_style='thin'),
                    bottom=Side(border_style='thin')
                )
                worksheet[f"C{n}"].border = Border(
                    left=Side(border_style='thin'),
                    right=Side(border_style='thin'),
                    top=Side(border_style='thin'),
                    bottom=Side(border_style='thin')
                )
                worksheet[f"D{n}"].border = Border(
                    left=Side(border_style='thin'),
                    right=Side(border_style='thin'),
                    top=Side(border_style='thin'),
                    bottom=Side(border_style='thin')
                )
                worksheet[f"E{n}"].border = Border(
                    left=Side(border_style='thin'),
                    right=Side(border_style='thin'),
                    top=Side(border_style='thin'),
                    bottom=Side(border_style='thin')
                )

            wb.save(tmp_path)
            wb["Sheet"].title = "Listado de siniestros"

            wb2 = wb.create_sheet("Estadisticas")

            wb2['A1'] = "Orquestador"
            wb2['A2'] = "Reporte Mensual de siniestros"
            wb2['A3'] = f"Reporte generado entre {f_inicio} al {f_fin}"
            wb2['A4'] = f"Reporte generado el {f_hoy} por - {email}"
            wb2["A1"].font = Font(bold=True)
            wb2["A2"].font = Font(bold=True)
            wb2["A3"].font = Font(bold=True)
            wb2["A4"].font = Font(bold=True)

            wb2['A6'] = 'DEPARTAMENTO'
            wb2["A6"].font = Font(bold=True)
            wb2['B6'] = 'MES'
            wb2["B6"].font = Font(bold=True)
            wb2['C6'] = 'AÑO'
            wb2["C6"].font = Font(bold=True)
            wb2['D6'] = 'EVALUADOS 1 VEZ'
            wb2["D6"].font = Font(bold=True)
            wb2['E6'] = '% EVALUADOS 1 VEZ'
            wb2["E6"].font = Font(bold=True)
            wb2['F6'] = 'EVALUADOS MÁS DE 1 VEZ'
            wb2["F6"].font = Font(bold=True)
            wb2['G6'] = '% EVALUADOS MÁS DE 1 VEZ'
            wb2["G6"].font = Font(bold=True)
            wb2['H6'] = 'FALLIDOS'
            wb2["H6"].font = Font(bold=True)
            wb2['I6'] = '% FALLIDOS'
            wb2["I6"].font = Font(bold=True)
            wb2['J6'] = 'TOTAL'
            wb2["J6"].font = Font(bold=True)

            wb2['A7'] = estadisticas["departamento"]
            wb2['B7'] = estadisticas["mes"]
            wb2['C7'] = estadisticas["anio"]
            wb2['D7'] = estadisticas["una_vez"]
            wb2['E7'] = estadisticas["porc_una_vez"]
            wb2['F7'] = estadisticas["varias_veces"]
            wb2['G7'] = estadisticas["porc_varias_veces"]
            wb2['H7'] = estadisticas["fallidos"]
            wb2['I7'] = estadisticas["porc_fallidos"]
            wb2['J7'] = estadisticas["total"]

            wb2.column_dimensions['A'].width = 16
            wb2.column_dimensions['D'].width = 17
            wb2.column_dimensions['E'].width = 19
            wb2.column_dimensions['F'].width = 25
            wb2.column_dimensions['G'].width = 26
            wb2.column_dimensions['H'].width = 10
            wb2.column_dimensions['I'].width = 11

            x = range(6, 8)
            for n in x:
                wb2[f"A{n}"].alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                    wrap_text=True
                )
                wb2[f"B{n}"].alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                    wrap_text=True
                )
                wb2[f"C{n}"].alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                    wrap_text=True
                )
                wb2[f"D{n}"].alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                    wrap_text=True
                )
                wb2[f"E{n}"].alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                    wrap_text=True
                )
                wb2[f"F{n}"].alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                    wrap_text=True
                )
                wb2[f"G{n}"].alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                    wrap_text=True
                )
                wb2[f"H{n}"].alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                    wrap_text=True
                )
                wb2[f"I{n}"].alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                    wrap_text=True
                )
                wb2[f"J{n}"].alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                    wrap_text=True
                )

                wb2[f"A{n}"].border = Border(
                    left=Side(border_style='thin'),
                    right=Side(border_style='thin'),
                    top=Side(border_style='thin'),
                    bottom=Side(border_style='thin')
                )
                wb2[f"B{n}"].border = Border(
                    left=Side(border_style='thin'),
                    right=Side(border_style='thin'),
                    top=Side(border_style='thin'),
                    bottom=Side(border_style='thin')
                )
                wb2[f"C{n}"].border = Border(
                    left=Side(border_style='thin'),
                    right=Side(border_style='thin'),
                    top=Side(border_style='thin'),
                    bottom=Side(border_style='thin')
                )
                wb2[f"D{n}"].border = Border(
                    left=Side(border_style='thin'),
                    right=Side(border_style='thin'),
                    top=Side(border_style='thin'),
                    bottom=Side(border_style='thin')
                )
                wb2[f"E{n}"].border = Border(
                    left=Side(border_style='thin'),
                    right=Side(border_style='thin'),
                    top=Side(border_style='thin'),
                    bottom=Side(border_style='thin')
                )
                wb2[f"F{n}"].border = Border(
                    left=Side(border_style='thin'),
                    right=Side(border_style='thin'),
                    top=Side(border_style='thin'),
                    bottom=Side(border_style='thin')
                )
                wb2[f"G{n}"].border = Border(
                    left=Side(border_style='thin'),
                    right=Side(border_style='thin'),
                    top=Side(border_style='thin'),
                    bottom=Side(border_style='thin')
                )
                wb2[f"H{n}"].border = Border(
                    left=Side(border_style='thin'),
                    right=Side(border_style='thin'),
                    top=Side(border_style='thin'),
                    bottom=Side(border_style='thin')
                )
                wb2[f"I{n}"].border = Border(
                    left=Side(border_style='thin'),
                    right=Side(border_style='thin'),
                    top=Side(border_style='thin'),
                    bottom=Side(border_style='thin')
                )
                wb2[f"J{n}"].border = Border(
                    left=Side(border_style='thin'),
                    right=Side(border_style='thin'),
                    top=Side(border_style='thin'),
                    bottom=Side(border_style='thin')
                )

            wb.save(tmp_path)

            secret = Aws.getSecret(str(os.getenv('STAGE'))+"/secret-orquestador")
            Aws.subirArchivoS3(secret['bucket_s3'], tmp_path, file_path)

            query = session.query(
                ReporteMensualSiniestros
            ).filter(
                ReporteMensualSiniestros.id == id_reporte
            )

            ##Vincula los nuevos datos que se van actualizar
            data = {
                'ruta': file_path,
                'respuesta': json.dumps(datos),
                'estado': 1
            }
            query.update(data)
            session.commit()
        
        except ValueError as e:
            raise Exception(f"Error al generar excel "+str(e))
        except Exception as e:
            raise Exception(f"Error al generar excel "+str(e))
        finally:
            session.invalidate()
            session.close()

    #Validamos si ya se generó el reporte
    def validarReporteMensual(self, id_reporte, id_usuario):
        db = Database('dbr').session
        try:
            resp_permission = self.validate_permission(user = id_usuario, type = "LISTAR")

            if resp_permission == False:
                raise ValueError("No cuenta con los permisos necesarios para realizar esta operación.")

            query = db.query(
                ReporteMensualSiniestros.id,
                ReporteMensualSiniestros.ruta,
                ReporteMensualSiniestros.respuesta,
                ReporteMensualSiniestros.estado,
            ).filter(
                ReporteMensualSiniestros.id == id_reporte,
                ReporteMensualSiniestros.estado.in_((1,2)),
            ).first()

            if query is None:
                msj = "Proceso en ejecución"
                return { "statusCode": 404, "message": msj}

            return { "statusCode": 200, "data": query }

        except ValueError as e:
            Type = type(e)
            raise Type(str(e))

        except Exception as error:
            return False
        
        finally:
            db.invalidate()
            db.close()

    # Funcion para valdiar el permiso que tiene un usuario
    @classmethod
    def validate_permission(cls, user: int, type: str, module: int = 10):
        try:
            data = {
                "USUARIO": user,
                "ID_MODULO": module,
                "PERMISO": type
            }

            resp_db = PermissionsDatabase.consult_permission(data)
            
            if resp_db is not None:
                return {'estado': True, "email": resp_db.CORREO}
            else:
                return False

        except Exception as e:
            raise ValueError(str(e))